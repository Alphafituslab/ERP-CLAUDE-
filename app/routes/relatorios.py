"""
Fase 7 — Painel Gerencial (BI básico): um único endpoint somente-leitura
que agrega indicadores das seis fases anteriores (Produção, Qualidade,
Estoque, Comercial, Financeiro) num só lugar, para quem precisa de visão
executiva sem navegar tela por tela.

Princípio de design: este módulo NÃO CRIA nenhum dado novo — tudo aqui é
`SELECT`/agregação sobre tabelas que já existem desde as fases 1 a 6. Por
isso não há migração de schema nesta fase. Cada número aqui é recalculado
a cada chamada a partir das mesmas fontes de verdade que as telas
operacionais já usam (ex.: saldo de estoque = soma de
`movimentacoes_estoque`, saldo de conta = `valor_total - SUM(baixas)`) —
nunca um valor pré-calculado e guardado à parte que poderia dessincronizar.

Por ser um resumo agregado (contagens e somas, não registros individuais
com detalhe operacional), a permissão `relatorios.visualizar` é
deliberadamente separada das permissões `*.visualizar` de cada módulo —
um perfil de Diretoria pode ter visão executiva completa sem precisar
receber acesso operacional a cada tela.
"""
import csv
import datetime
import io

from flask import Blueprint, Response, g, jsonify, request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .. import audit, painel_snapshot_service
from ..context import ApiError, client_device, client_ip, get_db
from ..pdf_marca import desenhar_cabecalho_logo
from ..permissions import requires_permission

bp = Blueprint("relatorios", __name__, url_prefix="/api/v1/relatorios")


def _hoje_iso_data():
    return datetime.datetime.utcnow().strftime("%Y-%m-%d")


def _empresa_ou_404(conn, empresa_id):
    """Fase 52 — mesma validação usada em app/routes/producao.py/lotes.py/
    comercial.py/financeiro.py, duplicada de propósito (ver comentário nos
    outros módulos)."""
    empresa = conn.execute("SELECT id, nome_fantasia, razao_social FROM empresas WHERE id = ?", (empresa_id,)).fetchone()
    if empresa is None:
        raise ApiError("Empresa não encontrada.", status=404)
    return dict(empresa)


def _contagem_por_status(conn, tabela, valores_possiveis, where_extra="", params=()):
    where = f"WHERE {where_extra}" if where_extra else ""
    rows = conn.execute(f"SELECT status, COUNT(*) AS total FROM {tabela} {where} GROUP BY status", params).fetchall()
    contagem = {v: 0 for v in valores_possiveis}
    for row in rows:
        if row["status"] in contagem:
            contagem[row["status"]] = row["total"]
    return contagem


def _bloco_producao(conn, empresa_id=None):
    """Fase 52 — `empresa_id` (opcional) filtra direto por
    `ordens_producao.empresa_id` (coluna nova, nullable — ver
    schema_fase52.sql). Sem o filtro, comportamento idêntico ao de antes
    desta fase."""
    where_extra = "empresa_id = ?" if empresa_id else ""
    params = (empresa_id,) if empresa_id else ()
    ordens_por_status = _contagem_por_status(
        conn, "ordens_producao", ("planejada", "liberada", "em_producao", "concluida", "cancelada"),
        where_extra, params,
    )
    query_lotes = "SELECT COUNT(*) AS total FROM ordens_producao WHERE status = 'concluida' AND lote_produzido_id IS NOT NULL"
    if empresa_id:
        query_lotes += " AND empresa_id = ?"
    lotes_produzidos = conn.execute(query_lotes, params).fetchone()["total"]
    return {"ordens_por_status": ordens_por_status, "lotes_produzidos_total": lotes_produzidos}


def _bloco_qualidade(conn, empresa_id=None):
    """Fase 52 — só `lotes_por_status`/`taxa_aprovacao_lotes_pct` são
    filtráveis por empresa (via `lotes.empresa_id`, coluna nova). `
    desvios_por_status` e `analises_aguardando_resultado` ficam
    DELIBERADAMENTE de fora do filtro: `desvios.lote_id` é nullable (nem
    todo desvio nasce de um lote recebido) e `analises` não tem nenhum
    vínculo com empresa — ver a nota completa em schema_fase52.sql."""
    where_extra = "empresa_id = ?" if empresa_id else ""
    params = (empresa_id,) if empresa_id else ()
    lotes_por_status = _contagem_por_status(
        conn, "lotes",
        ("quarentena", "em_analise", "aguardando_aprovacao", "aprovado",
         "aprovado_com_ressalva", "reprovado", "bloqueado", "devolvido", "destruido"),
        where_extra, params,
    )
    aprovados = lotes_por_status["aprovado"] + lotes_por_status["aprovado_com_ressalva"]
    reprovados = lotes_por_status["reprovado"]
    total_julgados = aprovados + reprovados
    taxa_aprovacao = round((aprovados / total_julgados) * 100, 1) if total_julgados > 0 else None

    desvios_por_status = _contagem_por_status(conn, "desvios", ("aberto", "em_tratativa", "encerrado"))
    analises_aguardando = conn.execute(
        "SELECT COUNT(*) AS total FROM analises WHERE status = 'aguardando_resultado'"
    ).fetchone()["total"]

    return {
        "lotes_por_status": lotes_por_status,
        "taxa_aprovacao_lotes_pct": taxa_aprovacao,
        "desvios_por_status": desvios_por_status,
        "analises_aguardando_resultado": analises_aguardando,
    }


def _bloco_estoque(conn, empresa_id=None):
    """Fase 52 — dois caminhos DELIBERADAMENTE diferentes para filtrar por
    empresa (ver a nota completa em schema_fase52.sql): saldo/vencidos
    usam a coluna nova `lotes.empresa_id` (mesma fonte de verdade da
    Qualidade, pra não ter dois números de "saldo do lote X" divergentes
    dentro do mesmo painel); já `posicoes_ativas` é uma métrica de
    INFRAESTRUTURA do armazém (a posição física em si — não o lote que
    está ocupando ela agora), então continua usando o caminho já existente
    desde a Fase 1 (posicoes_estoque → unidades → empresas)."""
    hoje = _hoje_iso_data()
    daqui_30_dias = (datetime.datetime.utcnow() + datetime.timedelta(days=30)).strftime("%Y-%m-%d")

    clausula_lote_empresa = " AND l.empresa_id = ?" if empresa_id else ""
    params_lote_empresa = (empresa_id,) if empresa_id else ()

    linhas_saldo = conn.execute(
        f"""
        SELECT i.tipo AS tipo, COALESCE(SUM(m.quantidade), 0) AS saldo
        FROM movimentacoes_estoque m
        JOIN lotes l ON l.id = m.lote_id
        JOIN itens i ON i.id = l.item_id
        WHERE 1=1{clausula_lote_empresa}
        GROUP BY i.tipo
        HAVING SUM(m.quantidade) > 0.0000001
        """,
        params_lote_empresa,
    ).fetchall()
    saldo_total_por_tipo_item = {row["tipo"]: row["saldo"] for row in linhas_saldo}

    # Lotes com saldo físico > 0 e vencidos (ou perto de vencer) — útil
    # para a Qualidade/PCP priorizar o que precisa de atenção, sem precisar
    # abrir a tela de Estoque lote a lote.
    lotes_com_saldo = conn.execute(
        f"""
        SELECT l.id, l.validade, COALESCE(SUM(m.quantidade), 0) AS saldo
        FROM lotes l JOIN movimentacoes_estoque m ON m.lote_id = l.id
        WHERE l.validade IS NOT NULL{clausula_lote_empresa}
        GROUP BY l.id
        HAVING SUM(m.quantidade) > 0.0000001
        """,
        params_lote_empresa,
    ).fetchall()
    lotes_vencidos = sum(1 for r in lotes_com_saldo if r["validade"] < hoje)
    lotes_a_vencer_30_dias = sum(1 for r in lotes_com_saldo if hoje <= r["validade"] <= daqui_30_dias)

    if empresa_id:
        posicoes_ativas = conn.execute(
            """
            SELECT COUNT(*) AS total FROM posicoes_estoque p
            JOIN unidades u ON u.id = p.unidade_id
            WHERE p.status = 'ativa' AND u.empresa_id = ?
            """,
            (empresa_id,),
        ).fetchone()["total"]
    else:
        posicoes_ativas = conn.execute(
            "SELECT COUNT(*) AS total FROM posicoes_estoque WHERE status = 'ativa'"
        ).fetchone()["total"]

    return {
        "saldo_total_por_tipo_item": saldo_total_por_tipo_item,
        "posicoes_ativas": posicoes_ativas,
        "lotes_vencidos_com_saldo": lotes_vencidos,
        "lotes_a_vencer_30_dias_com_saldo": lotes_a_vencer_30_dias,
    }


def _bloco_comercial(conn, empresa_id=None):
    """Fase 52 — `pedidos_por_status`/`valor_total_expedido` filtram via
    `pedidos_venda.empresa_id` (coluna nova). `clientes_ativos` fica
    DELIBERADAMENTE de fora do filtro: um cliente não tem vínculo natural
    com uma empresa específica do grupo — ver schema_fase52.sql."""
    where_extra = "empresa_id = ?" if empresa_id else ""
    params = (empresa_id,) if empresa_id else ()
    pedidos_por_status = _contagem_por_status(
        conn, "pedidos_venda", ("rascunho", "confirmado", "expedido", "cancelado"), where_extra, params
    )
    clausula_empresa_pv = " AND pv.empresa_id = ?" if empresa_id else ""
    valor_total_expedido = conn.execute(
        f"""
        SELECT COALESCE(SUM(pvi.quantidade * pvi.preco_unitario), 0) AS total
        FROM pedido_venda_itens pvi
        JOIN pedidos_venda pv ON pv.id = pvi.pedido_id
        WHERE pv.status = 'expedido'{clausula_empresa_pv}
        """,
        params,
    ).fetchone()["total"]
    clientes_ativos = conn.execute("SELECT COUNT(*) AS total FROM clientes WHERE status = 'ativo'").fetchone()["total"]

    return {
        "pedidos_por_status": pedidos_por_status,
        "valor_total_expedido": valor_total_expedido,
        "clientes_ativos": clientes_ativos,
    }


def _baixado_liquido(conn, tabela_baixas, campo_fk, conta_id):
    """Soma as baixas normais e SUBTRAI as baixas de estorno (Fase 14) de
    uma conta — mesma lógica de `_total_baixado()` em
    app/routes/financeiro.py, centralizada aqui para as duas funções deste
    módulo que precisam dela (`_bloco_financeiro` e
    `fluxo_caixa_projetado`) nunca divergirem de novo uma da outra. Uma
    baixa de estorno sempre tem o MESMO valor da baixa original que ela
    neutraliza, então a subtração deixa o saldo em aberto exatamente como
    se a baixa original nunca tivesse existido — sem jamais dar
    UPDATE/DELETE na linha original (bloqueado por trigger desde a Fase 6).
    (Esta duplicação de lógica entre `financeiro.py` e este módulo já
    causou uma regressão real: o painel gerencial ficou temporariamente
    contando estornos como se fossem baixas normais, até ser corrigido.)"""
    return conn.execute(
        f"""
        SELECT
            COALESCE(SUM(CASE WHEN estorno_de_id IS NULL THEN valor ELSE 0 END), 0)
            - COALESCE(SUM(CASE WHEN estorno_de_id IS NOT NULL THEN valor ELSE 0 END), 0) AS total
        FROM {tabela_baixas} WHERE {campo_fk} = ?
        """,
        (conta_id,),
    ).fetchone()["total"]


def _contas_em_aberto(conn, tabela_conta, tabela_baixas, campo_fk, empresa_id=None):
    """Lista de dicts {id, valor_total, vencimento, saldo_aberto} só para
    contas com status 'aberto'/'pago_parcial' — saldo_aberto sempre
    recalculado a partir do ledger completo (nunca um número guardado à
    parte), usada tanto pelo resumo do dashboard quanto pelo fluxo de
    caixa projetado (Fase 15).

    Fase 52 — `empresa_id` (opcional) filtra direto por
    `{tabela_conta}.empresa_id` (coluna nova em contas_receber e
    contas_pagar — ver schema_fase52.sql). `contas_receber.empresa_id` é
    herdada automaticamente do pedido de venda na expedição
    (comercial.py:expedir); `contas_pagar.empresa_id` é informada na
    criação (financeiro.py:criar_conta_pagar)."""
    query = f"SELECT id, valor_total, vencimento FROM {tabela_conta} WHERE status IN ('aberto', 'pago_parcial')"
    params = ()
    if empresa_id:
        query += " AND empresa_id = ?"
        params = (empresa_id,)
    contas = conn.execute(query, params).fetchall()
    resultado = []
    for conta in contas:
        baixado = _baixado_liquido(conn, tabela_baixas, campo_fk, conta["id"])
        resultado.append({
            "id": conta["id"],
            "valor_total": conta["valor_total"],
            "vencimento": conta["vencimento"],
            "saldo_aberto": conta["valor_total"] - baixado,
        })
    return resultado


def _bloco_financeiro(conn, empresa_id=None):
    """Fase 52 — `empresa_id` (opcional) filtra `contas_receber`/
    `contas_pagar` pela coluna nova `empresa_id` de cada tabela."""
    hoje = _hoje_iso_data()

    def _resumo(tabela_conta, tabela_baixas, campo_fk):
        total_recebido_pago = 0.0
        # total_baixado soma TODAS as contas (inclusive já pagas/canceladas
        # com baixa histórica), então não dá para reaproveitar
        # `_contas_em_aberto` aqui — ela já filtra só as em aberto.
        query_todas = f"SELECT id FROM {tabela_conta}"
        params_todas = ()
        if empresa_id:
            query_todas += " WHERE empresa_id = ?"
            params_todas = (empresa_id,)
        todas = conn.execute(query_todas, params_todas).fetchall()
        for conta in todas:
            total_recebido_pago += _baixado_liquido(conn, tabela_baixas, campo_fk, conta["id"])

        total_aberto = 0.0
        total_vencido = 0.0
        for conta in _contas_em_aberto(conn, tabela_conta, tabela_baixas, campo_fk, empresa_id):
            total_aberto += conta["saldo_aberto"]
            if conta["vencimento"] < hoje:
                total_vencido += conta["saldo_aberto"]

        return {
            "total_em_aberto": round(total_aberto, 2),
            "total_vencido": round(total_vencido, 2),
            "total_baixado": round(total_recebido_pago, 2),
        }

    receber = _resumo("contas_receber", "contas_receber_baixas", "conta_receber_id")
    pagar = _resumo("contas_pagar", "contas_pagar_baixas", "conta_pagar_id")
    saldo_projetado = round(receber["total_em_aberto"] - pagar["total_em_aberto"], 2)

    return {"contas_a_receber": receber, "contas_a_pagar": pagar, "saldo_projetado_em_aberto": saldo_projetado}


# ============================================================
# Fase 15 — Fluxo de Caixa Projetado
# ============================================================
# Agrupa o saldo em aberto de contas a receber/pagar por faixa de dias até
# o vencimento (o mesmo "aging" usado em qualquer AR/AP), com um saldo
# acumulado faixa a faixa — responde "como fica o caixa daqui a X dias se
# nada mais entrar/sair além do que já está lançado?". 100% derivado a
# cada chamada a partir de `_contas_em_aberto` acima (mesma fonte de
# verdade do bloco "financeiro" do dashboard) — nenhuma tabela nova,
# nenhum valor pré-calculado guardado à parte.
_FLUXO_CAIXA_BUCKETS = (
    ("vencido", "Vencido", None, -1),
    ("0_7", "0 a 7 dias", 0, 7),
    ("8_15", "8 a 15 dias", 8, 15),
    ("16_30", "16 a 30 dias", 16, 30),
    ("31_60", "31 a 60 dias", 31, 60),
    ("61_90", "61 a 90 dias", 61, 90),
    ("mais_90", "Mais de 90 dias", 91, None),
)


def _bucket_por_dias(dias_ate_vencimento):
    for chave, _rotulo, minimo, maximo in _FLUXO_CAIXA_BUCKETS:
        if minimo is not None and dias_ate_vencimento < minimo:
            continue
        if maximo is not None and dias_ate_vencimento > maximo:
            continue
        return chave
    return "mais_90"  # nunca deveria cair aqui — último bucket já é aberto (maximo=None)


def _fluxo_caixa_projetado(conn, empresa_id=None):
    """Fase 52 — `empresa_id` (opcional) propaga para `_contas_em_aberto`,
    então o fluxo de caixa projetado fica coerente com o filtro aplicado
    no resto do Painel Gerencial (mesma fonte de verdade)."""
    hoje = datetime.datetime.utcnow().date()
    somas = {chave: {"entradas_previstas": 0.0, "saidas_previstas": 0.0} for chave, *_ in _FLUXO_CAIXA_BUCKETS}

    def _acumular(contas, campo):
        for conta in contas:
            vencimento = datetime.datetime.strptime(conta["vencimento"], "%Y-%m-%d").date()
            dias = (vencimento - hoje).days
            chave = _bucket_por_dias(dias)
            somas[chave][campo] += conta["saldo_aberto"]

    _acumular(_contas_em_aberto(conn, "contas_receber", "contas_receber_baixas", "conta_receber_id", empresa_id), "entradas_previstas")
    _acumular(_contas_em_aberto(conn, "contas_pagar", "contas_pagar_baixas", "conta_pagar_id", empresa_id), "saidas_previstas")

    buckets = []
    saldo_acumulado = 0.0
    for chave, rotulo, _minimo, _maximo in _FLUXO_CAIXA_BUCKETS:
        entradas = round(somas[chave]["entradas_previstas"], 2)
        saidas = round(somas[chave]["saidas_previstas"], 2)
        saldo_liquido = round(entradas - saidas, 2)
        saldo_acumulado = round(saldo_acumulado + saldo_liquido, 2)
        buckets.append({
            "bucket": chave,
            "rotulo": rotulo,
            "entradas_previstas": entradas,
            "saidas_previstas": saidas,
            "saldo_liquido": saldo_liquido,
            "saldo_acumulado": saldo_acumulado,
        })

    return {"hoje": hoje.isoformat(), "buckets": buckets}


@bp.get("/fluxo-caixa-projetado")
@requires_permission("relatorios", "visualizar")
def fluxo_caixa_projetado():
    """`empresa_id` (opcional, Fase 52) filtra o fluxo de caixa pela mesma
    coluna nova usada no resto do Painel Gerencial."""
    conn = get_db()
    empresa_id = request.args.get("empresa_id", type=int)
    if empresa_id:
        _empresa_ou_404(conn, empresa_id)
    return jsonify(_fluxo_caixa_projetado(conn, empresa_id))


# ============================================================
# Fase 42 — Painel Gerencial: Filtro por Período
# ============================================================
# Os cinco blocos acima (Produção/Qualidade/Estoque/Comercial/Financeiro)
# sempre mostraram a "situação atual" — uma foto de agora, sem filtro —
# desde a Fase 7. Isso continua assim de propósito: saldo de estoque e
# contas em aberto são, por natureza, o estado ATUAL das coisas (não faz
# sentido "o saldo de estoque de 1º de janeiro", só o saldo de agora).
# O que este bloco adiciona é um sexto bloco, OPCIONAL e aditivo — nunca
# substitui nem muda os cinco de cima — com indicadores de FLUXO (o que
# aconteceu dentro de uma janela de tempo), no mesmo espírito de
# `data_inicio`/`data_fim` que o DRE já usa desde a Fase 20/41
# (`custeio._dre_simplificado`). Sem os parâmetros, o dashboard continua
# devolvendo exatamente o que devolvia antes desta fase — nenhuma
# consulta nova roda, `periodo.aplicado` vem `false`.
def _bloco_periodo(conn, data_inicio, data_fim, empresa_id=None):
    """Fase 52 — `empresa_id` (opcional) COMPÕE com o filtro de período
    (não o substitui): quando os dois estão presentes, o bloco mostra o
    que aconteceu no período E dentro daquela empresa. Ordens/lotes/
    pedidos filtram direto pela própria coluna `empresa_id`; as baixas
    (contas_receber_baixas/contas_pagar_baixas) não têm `empresa_id`
    própria (é a CONTA que tem — ver schema_fase52.sql), então precisam de
    um JOIN até a tabela da conta pai para aplicar o filtro."""
    if not data_inicio and not data_fim:
        return {"aplicado": False, "data_inicio": None, "data_fim": None}

    def _clausula(coluna):
        clausulas, params = [], []
        if data_inicio:
            clausulas.append(f"{coluna} >= ?")
            params.append(data_inicio)
        if data_fim:
            clausulas.append(f"{coluna} <= ?")
            params.append(f"{data_fim}T23:59:59.999999Z")
        return clausulas, params

    # Produção — ordens concluídas no período, por `concluido_em`.
    clausulas, params = _clausula("concluido_em")
    if empresa_id:
        clausulas.append("empresa_id = ?")
        params.append(empresa_id)
    where = " AND ".join(["status = 'concluida'"] + clausulas)
    ordens_concluidas_periodo = conn.execute(
        f"SELECT COUNT(*) AS total FROM ordens_producao WHERE {where}", params
    ).fetchone()["total"]

    # Qualidade — lotes aprovados/reprovados no período. Simplificação
    # documentada (mesmo espírito da Fase 13/34): usa `atualizado_em`
    # como aproximação de "quando o status virou aprovado/reprovado",
    # porque não existe uma coluna própria "data_aprovacao"/
    # "data_reprovacao" na tabela `lotes`. Se um lote já aprovado for
    # editado por outro motivo depois (ex.: corrigir a validade), o
    # `atualizado_em` avança e ele pode aparecer no período errado — um
    # caso raro, mas real, que fica documentado aqui.
    clausulas, params = _clausula("atualizado_em")
    if empresa_id:
        clausulas.append("empresa_id = ?")
        params.append(empresa_id)
    where_aprovados = " AND ".join(["status IN ('aprovado', 'aprovado_com_ressalva')"] + clausulas)
    where_reprovados = " AND ".join(["status = 'reprovado'"] + clausulas)
    lotes_aprovados_periodo = conn.execute(f"SELECT COUNT(*) AS total FROM lotes WHERE {where_aprovados}", params).fetchone()["total"]
    lotes_reprovados_periodo = conn.execute(f"SELECT COUNT(*) AS total FROM lotes WHERE {where_reprovados}", params).fetchone()["total"]

    # Comercial — pedidos expedidos e valor expedido no período, mesma
    # coluna/regime de competência já usado pelo DRE (`pv.expedido_em`).
    clausulas, params = _clausula("pv.expedido_em")
    if empresa_id:
        clausulas.append("pv.empresa_id = ?")
        params.append(empresa_id)
    where = " AND ".join(["pv.status = 'expedido'"] + clausulas)
    pedidos_expedidos_periodo = conn.execute(
        f"SELECT COUNT(*) AS total FROM pedidos_venda pv WHERE {where}", params
    ).fetchone()["total"]
    valor_expedido_periodo = conn.execute(
        f"""
        SELECT COALESCE(SUM(pvi.quantidade * pvi.preco_unitario), 0) AS total
        FROM pedido_venda_itens pvi
        JOIN pedidos_venda pv ON pv.id = pvi.pedido_id
        WHERE {where}
        """,
        params,
    ).fetchone()["total"]

    # Financeiro — baixado líquido (baixas normais − estornos, mesma
    # regra de `_baixado_liquido`) no período, por `criado_em` da baixa
    # (quando o recebimento/pagamento foi de fato registrado). A baixa em
    # si não tem `empresa_id` — precisa de um JOIN até a conta pai para
    # aplicar o filtro por empresa (Fase 52).
    def _baixado_periodo(tabela_baixas, tabela_conta, campo_fk):
        clausulas, params = _clausula("b.criado_em")
        where = " AND ".join(clausulas) if clausulas else "1=1"
        join_empresa = ""
        if empresa_id:
            join_empresa = f" JOIN {tabela_conta} c ON c.id = b.{campo_fk}"
            where = " AND ".join([where, "c.empresa_id = ?"])
            params = params + [empresa_id]
        return conn.execute(
            f"""
            SELECT
                COALESCE(SUM(CASE WHEN b.estorno_de_id IS NULL THEN b.valor ELSE 0 END), 0)
                - COALESCE(SUM(CASE WHEN b.estorno_de_id IS NOT NULL THEN b.valor ELSE 0 END), 0) AS total
            FROM {tabela_baixas} b{join_empresa} WHERE {where}
            """,
            params,
        ).fetchone()["total"]

    valor_recebido_periodo = _baixado_periodo("contas_receber_baixas", "contas_receber", "conta_receber_id")
    valor_pago_periodo = _baixado_periodo("contas_pagar_baixas", "contas_pagar", "conta_pagar_id")

    return {
        "aplicado": True,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "ordens_concluidas_no_periodo": ordens_concluidas_periodo,
        "lotes_aprovados_no_periodo": lotes_aprovados_periodo,
        "lotes_reprovados_no_periodo": lotes_reprovados_periodo,
        "pedidos_expedidos_no_periodo": pedidos_expedidos_periodo,
        "valor_expedido_no_periodo": round(valor_expedido_periodo, 2),
        "valor_recebido_no_periodo": round(valor_recebido_periodo, 2),
        "valor_pago_no_periodo": round(valor_pago_periodo, 2),
    }


def _montar_dashboard(conn, data_inicio=None, data_fim=None, empresa_id=None):
    """Monta o mesmo dict que `GET /dashboard` devolve — extraído para uma
    função só (em vez de repetir o `jsonify({...})` inline) para a rota de
    exportação em PDF (Fase 18) reaproveitar exatamente a mesma agregação,
    nunca correndo o risco de o PDF mostrar um número calculado de um jeito
    sutilmente diferente do que a tela já mostra (mesmo motivo pelo qual
    `_contas_em_aberto`/`_baixado_liquido` foram centralizadas na Fase 15).
    `data_inicio`/`data_fim` (Fase 42, opcionais) só afetam o bloco
    `periodo` — os cinco blocos "situação atual" nunca mudam com eles.

    Fase 52 — `empresa_id` (opcional) propaga para TODOS os blocos (ao
    contrário do período, que só afeta o bloco `periodo`), já que "ver a
    situação atual só da empresa X" é o próprio propósito do filtro. Sem
    `empresa_id`, cada bloco roda exatamente a mesma consulta de antes
    desta fase — comportamento 100% retrocompatível."""
    return {
        "gerado_em": datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
        "empresa_filtrada": _empresa_ou_404(conn, empresa_id) if empresa_id else None,
        "producao": _bloco_producao(conn, empresa_id),
        "qualidade": _bloco_qualidade(conn, empresa_id),
        "estoque": _bloco_estoque(conn, empresa_id),
        "comercial": _bloco_comercial(conn, empresa_id),
        "financeiro": _bloco_financeiro(conn, empresa_id),
        "periodo": _bloco_periodo(conn, data_inicio, data_fim, empresa_id),
    }


@bp.get("/dashboard")
@requires_permission("relatorios", "visualizar")
def dashboard():
    """`data_inicio`/`data_fim` (AAAA-MM-DD, opcionais, Fase 42) filtram só
    o bloco `periodo` do retorno. `empresa_id` (opcional, Fase 52) filtra
    os cinco blocos de "situação atual" inteiros — os dois filtros são
    independentes e compõem entre si."""
    conn = get_db()
    data_inicio = request.args.get("data_inicio") or None
    data_fim = request.args.get("data_fim") or None
    empresa_id = request.args.get("empresa_id", type=int)
    if empresa_id:
        _empresa_ou_404(conn, empresa_id)
    dados = _montar_dashboard(conn, data_inicio, data_fim, empresa_id)
    # Fase 69 — efeito colateral barato desta mesma chamada: grava/atualiza
    # o snapshot de HOJE para a série histórica (ver a nota de escopo
    # completa em app/painel_snapshot_service.py e migrations/schema_fase69.sql).
    # Nunca falha a rota por causa disso — é um adicional "melhor esforço",
    # a tela continua funcionando normalmente mesmo se este passo falhar.
    try:
        painel_snapshot_service.capturar_ou_atualizar_snapshot_do_dia(conn, dados, empresa_id)
    except Exception:
        pass
    return jsonify(dados)


# ============================================================
# Fase 69 — Painel Gerencial: Série Histórica / Tendência
# ============================================================
@bp.get("/painel/tendencia")
@requires_permission("relatorios", "visualizar")
def painel_tendencia():
    """Devolve a série histórica capturada pelos snapshots diários (ver
    painel_snapshot_service.listar_tendencia) — mesma permissão do resto
    do Painel Gerencial, nenhuma permissão nova. `dias` (opcional, padrão
    30, máximo 365) e `empresa_id` (opcional, mesmo padrão da Fase 52)."""
    conn = get_db()
    dias = request.args.get("dias", default=painel_snapshot_service.DIAS_PADRAO, type=int)
    empresa_id = request.args.get("empresa_id", type=int)
    if empresa_id:
        _empresa_ou_404(conn, empresa_id)
    return jsonify(painel_snapshot_service.listar_tendencia(conn, dias, empresa_id))


# ============================================================
# Fase 18 — Exportação do Painel Gerencial em PDF
# ============================================================
# Mesmo padrão da Fase 10 (CoA) e Fase 11 (Relatório de Recall): um export
# puro, sem alterar nenhum dado de negócio, formatando com reportlab
# (platypus) exatamente os mesmos números que `GET /dashboard` e
# `GET /fluxo-caixa-projetado` já servem para a tela — nenhuma tabela nova,
# nenhum valor pré-calculado guardado à parte. Reaproveita a mesma
# permissão `relatorios.visualizar` do resto do Painel Gerencial (nenhuma
# permissão nova), porque exportar é exatamente a mesma capacidade de
# "visualizar", só num formato diferente.
_ROTULOS_TIPO_ITEM = {
    "materia_prima": "Matéria-prima",
    "embalagem_primaria": "Embalagem primária",
    "embalagem_secundaria": "Embalagem secundária",
    "produto_intermediario": "Produto intermediário",
    "produto_a_granel": "Produto a granel",
    "produto_acabado": "Produto acabado",
    "material_de_laboratorio": "Material de laboratório",
}


def _gerar_pdf_dashboard(dados, fluxo_caixa, gerado_por_nome):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=1.8 * cm, bottomMargin=1.8 * cm, leftMargin=2 * cm, rightMargin=2 * cm,
    )
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "TituloPainel", parent=estilos["Title"], fontSize=16, alignment=TA_CENTER, spaceAfter=2,
    )
    estilo_subtitulo = ParagraphStyle(
        "SubtituloPainel", parent=estilos["Normal"], fontSize=10, alignment=TA_CENTER,
        textColor=colors.HexColor("#5a6472"), spaceAfter=16,
    )
    estilo_rotulo_secao = ParagraphStyle(
        "RotuloSecaoPainel", parent=estilos["Heading3"], fontSize=11, spaceBefore=12, spaceAfter=6,
        textColor=colors.HexColor("#1f3a5f"),
    )

    def _tabela_kpi(linhas, larguras=(9 * cm, 7 * cm)):
        tabela = Table(linhas, colWidths=list(larguras))
        tabela.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e6ea")),
        ]))
        return tabela

    def _moeda(valor):
        return f"R$ {valor:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")

    elementos = [
        Paragraph("Alphafitus Laboratório Nutracêutico Ltda.", estilo_titulo),
        Paragraph(
            f"Painel Gerencial — gerado em {dados['gerado_em']} por {gerado_por_nome}",
            estilo_subtitulo,
        ),
    ]
    # Fase 52 — quando o filtro por empresa foi aplicado, deixa isso
    # explícito no cabeçalho do export, senão o número fica sem contexto
    # de qual empresa ele representa.
    empresa_filtrada = dados.get("empresa_filtrada")
    if empresa_filtrada:
        nome_empresa = empresa_filtrada["nome_fantasia"] or empresa_filtrada["razao_social"]
        elementos.append(Paragraph(f"Filtrando por: {nome_empresa}", estilo_subtitulo))

    op = dados["producao"]["ordens_por_status"]
    elementos.append(Paragraph("Produção", estilo_rotulo_secao))
    elementos.append(_tabela_kpi([
        ["Planejadas", str(op["planejada"])],
        ["Liberadas", str(op["liberada"])],
        ["Em produção", str(op["em_producao"])],
        ["Concluídas", str(op["concluida"])],
        ["Canceladas", str(op["cancelada"])],
        ["Lotes produzidos (total)", str(dados["producao"]["lotes_produzidos_total"])],
    ]))

    lp = dados["qualidade"]["lotes_por_status"]
    ds = dados["qualidade"]["desvios_por_status"]
    taxa = dados["qualidade"]["taxa_aprovacao_lotes_pct"]
    elementos.append(Paragraph("Qualidade", estilo_rotulo_secao))
    elementos.append(_tabela_kpi([
        ["Aprovados", str(lp["aprovado"] + lp["aprovado_com_ressalva"])],
        ["Reprovados", str(lp["reprovado"])],
        ["Em quarentena/análise", str(lp["quarentena"] + lp["em_analise"] + lp["aguardando_aprovacao"])],
        ["Taxa de aprovação", f"{taxa}%" if taxa is not None else "—"],
        ["Análises aguardando resultado", str(dados["qualidade"]["analises_aguardando_resultado"])],
        ["Desvios em aberto", f'{ds["aberto"]} ({ds["em_tratativa"]} em tratativa)'],
    ]))

    est = dados["estoque"]
    elementos.append(Paragraph("Estoque (WMS)", estilo_rotulo_secao))
    linhas_estoque = [
        ["Posições ativas", str(est["posicoes_ativas"])],
        ["Lotes vencidos com saldo", str(est["lotes_vencidos_com_saldo"])],
        ["A vencer em 30 dias", str(est["lotes_a_vencer_30_dias_com_saldo"])],
    ]
    for tipo, saldo in est["saldo_total_por_tipo_item"].items():
        linhas_estoque.append([f"Saldo — {_ROTULOS_TIPO_ITEM.get(tipo, tipo)}", f"{saldo:.3f}"])
    elementos.append(_tabela_kpi(linhas_estoque))

    cp = dados["comercial"]["pedidos_por_status"]
    elementos.append(Paragraph("Comercial", estilo_rotulo_secao))
    elementos.append(_tabela_kpi([
        ["Rascunho", str(cp["rascunho"])],
        ["Confirmados", str(cp["confirmado"])],
        ["Expedidos", str(cp["expedido"])],
        ["Cancelados", str(cp["cancelado"])],
        ["Valor total expedido", _moeda(dados["comercial"]["valor_total_expedido"])],
        ["Clientes ativos", str(dados["comercial"]["clientes_ativos"])],
    ]))

    fr = dados["financeiro"]
    elementos.append(Paragraph("Financeiro", estilo_rotulo_secao))
    elementos.append(_tabela_kpi([
        ["A receber em aberto", _moeda(fr["contas_a_receber"]["total_em_aberto"])],
        ["  dos quais vencido", _moeda(fr["contas_a_receber"]["total_vencido"])],
        ["Recebido (total)", _moeda(fr["contas_a_receber"]["total_baixado"])],
        ["A pagar em aberto", _moeda(fr["contas_a_pagar"]["total_em_aberto"])],
        ["  dos quais vencido", _moeda(fr["contas_a_pagar"]["total_vencido"])],
        ["Pago (total)", _moeda(fr["contas_a_pagar"]["total_baixado"])],
        ["Saldo projetado (aberto)", _moeda(fr["saldo_projetado_em_aberto"])],
    ]))

    elementos.append(Paragraph(f"Fluxo de Caixa Projetado (calculado em {fluxo_caixa['hoje']})", estilo_rotulo_secao))
    cabecalho = ["Faixa", "Entradas previstas", "Saídas previstas", "Saldo líquido", "Saldo acumulado"]
    linhas_fluxo = [cabecalho] + [
        [b["rotulo"], _moeda(b["entradas_previstas"]), _moeda(b["saidas_previstas"]),
         _moeda(b["saldo_liquido"]), _moeda(b["saldo_acumulado"])]
        for b in fluxo_caixa["buckets"]
    ]
    tabela_fluxo = Table(linhas_fluxo, colWidths=[3.2 * cm, 3.4 * cm, 3.4 * cm, 3 * cm, 3 * cm])
    tabela_fluxo.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e6ea")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
    ]))
    elementos.append(tabela_fluxo)

    # Fase 42 — bloco extra "no período", só quando o filtro foi usado.
    periodo = dados.get("periodo") or {"aplicado": False}
    if periodo["aplicado"]:
        elementos.append(Paragraph(
            f"No período — de {periodo['data_inicio'] or '(sem início)'} até {periodo['data_fim'] or '(sem fim)'}",
            estilo_rotulo_secao,
        ))
        elementos.append(_tabela_kpi([
            ["Ordens concluídas", str(periodo["ordens_concluidas_no_periodo"])],
            ["Lotes aprovados", str(periodo["lotes_aprovados_no_periodo"])],
            ["Lotes reprovados", str(periodo["lotes_reprovados_no_periodo"])],
            ["Pedidos expedidos", str(periodo["pedidos_expedidos_no_periodo"])],
            ["Valor expedido", _moeda(periodo["valor_expedido_no_periodo"])],
            ["Valor recebido", _moeda(periodo["valor_recebido_no_periodo"])],
            ["Valor pago", _moeda(periodo["valor_pago_no_periodo"])],
        ]))

    elementos.append(Spacer(1, 0.4 * cm))
    elementos.append(Paragraph(
        "Documento gerado automaticamente a partir dos dados operacionais no momento da exportação — "
        "não substitui os relatórios contábeis/fiscais oficiais.",
        ParagraphStyle("RodapePainel", parent=estilos["Normal"], fontSize=7.5, textColor=colors.HexColor("#8a92a0")),
    ))

    doc.build(elementos, onFirstPage=desenhar_cabecalho_logo, onLaterPages=desenhar_cabecalho_logo)
    return buffer.getvalue()


@bp.get("/dashboard/pdf")
@requires_permission("relatorios", "visualizar")
def dashboard_pdf():
    """Fase 18 — exporta o Painel Gerencial (os mesmos 5 blocos de
    `GET /dashboard` + o Fluxo de Caixa Projetado da Fase 15) como um
    documento PDF, para quem precisa levar os números pra uma reunião ou
    arquivar um snapshot do dia. Export puro — não altera nenhum dado de
    negócio. Grava um evento de auditoria com `tabela="painel_gerencial"`,
    um rótulo sintético (não uma tabela real do banco, já que o painel é
    100% agregação) só para deixar rastreável quem exportou o quê e
    quando, seguindo o mesmo princípio de auditoria de toda ação relevante
    das fases anteriores."""
    usuario_atual = g.usuario_atual
    conn = get_db()
    data_inicio = request.args.get("data_inicio") or None
    data_fim = request.args.get("data_fim") or None
    empresa_id = request.args.get("empresa_id", type=int)
    if empresa_id:
        _empresa_ou_404(conn, empresa_id)
    dados = _montar_dashboard(conn, data_inicio, data_fim, empresa_id)
    fluxo_caixa = _fluxo_caixa_projetado(conn, empresa_id)
    pdf_bytes = _gerar_pdf_dashboard(dados, fluxo_caixa, usuario_atual["nome"])

    audit.registrar(
        conn, tabela="painel_gerencial", registro_id=None, usuario_id=usuario_atual["id"],
        acao="painel_pdf_gerado", valor_novo={"gerado_em": dados["gerado_em"]},
        ip=client_ip(), dispositivo=client_device(),
    )

    data_arquivo = datetime.datetime.utcnow().strftime("%Y-%m-%d")
    nome_arquivo = f"Painel-Gerencial-{data_arquivo}.pdf"
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


# ============================================================
# Fase 19 — Exportação do Painel Gerencial em CSV
# ============================================================
# Mesmo princípio da Fase 18 (PDF), mas para quem quer os números crus
# numa planilha em vez de um documento formatado — reaproveita exatamente
# a mesma `_montar_dashboard`/`_fluxo_caixa_projetado` da Fase 18, então o
# CSV e o PDF (e a própria tela) nunca podem divergir um do outro. Só a
# biblioteca padrão do Python (`csv`), nenhuma dependência nova: ao
# contrário de um export em .xlsx de verdade (que precisaria de uma
# biblioteca como openpyxl), um CSV abre direto no Excel/LibreOffice sem
# nenhuma dependência extra, mantendo o backend com a mesma pegada mínima
# de sempre.
def _gerar_csv_dashboard(dados, fluxo_caixa):
    # `﻿` (BOM) no início: sem isso, o Excel no Windows abre um CSV
    # UTF-8 com acentos quebrados (ex.: "Produção" vira "ProduÃ§Ã£o") — é a
    # mesma dor de cabeça clássica de qualquer CSV em português aberto no
    # Excel, então incluímos o BOM de propósito para abrir corretamente
    # sem o usuário precisar saber importar como UTF-8 manualmente.
    buffer = io.StringIO()
    buffer.write("﻿")
    w = csv.writer(buffer, delimiter=";")

    w.writerow([f"Painel Gerencial — gerado em {dados['gerado_em']}"])
    # Fase 52 — mesmo motivo do PDF: deixar explícito qual empresa os
    # números representam quando o filtro foi aplicado.
    empresa_filtrada = dados.get("empresa_filtrada")
    if empresa_filtrada:
        nome_empresa = empresa_filtrada["nome_fantasia"] or empresa_filtrada["razao_social"]
        w.writerow([f"Filtrando por: {nome_empresa}"])
    w.writerow([])

    op = dados["producao"]["ordens_por_status"]
    w.writerow(["Produção"])
    w.writerow(["Indicador", "Valor"])
    w.writerow(["Planejadas", op["planejada"]])
    w.writerow(["Liberadas", op["liberada"]])
    w.writerow(["Em produção", op["em_producao"]])
    w.writerow(["Concluídas", op["concluida"]])
    w.writerow(["Canceladas", op["cancelada"]])
    w.writerow(["Lotes produzidos (total)", dados["producao"]["lotes_produzidos_total"]])
    w.writerow([])

    lp = dados["qualidade"]["lotes_por_status"]
    ds = dados["qualidade"]["desvios_por_status"]
    taxa = dados["qualidade"]["taxa_aprovacao_lotes_pct"]
    w.writerow(["Qualidade"])
    w.writerow(["Indicador", "Valor"])
    w.writerow(["Aprovados", lp["aprovado"] + lp["aprovado_com_ressalva"]])
    w.writerow(["Reprovados", lp["reprovado"]])
    w.writerow(["Em quarentena/análise", lp["quarentena"] + lp["em_analise"] + lp["aguardando_aprovacao"]])
    w.writerow(["Taxa de aprovação (%)", taxa if taxa is not None else ""])
    w.writerow(["Análises aguardando resultado", dados["qualidade"]["analises_aguardando_resultado"]])
    w.writerow(["Desvios em aberto", ds["aberto"]])
    w.writerow(["Desvios em tratativa", ds["em_tratativa"]])
    w.writerow([])

    est = dados["estoque"]
    w.writerow(["Estoque (WMS)"])
    w.writerow(["Indicador", "Valor"])
    w.writerow(["Posições ativas", est["posicoes_ativas"]])
    w.writerow(["Lotes vencidos com saldo", est["lotes_vencidos_com_saldo"]])
    w.writerow(["A vencer em 30 dias", est["lotes_a_vencer_30_dias_com_saldo"]])
    for tipo, saldo in est["saldo_total_por_tipo_item"].items():
        w.writerow([f"Saldo — {_ROTULOS_TIPO_ITEM.get(tipo, tipo)}", saldo])
    w.writerow([])

    cp = dados["comercial"]["pedidos_por_status"]
    w.writerow(["Comercial"])
    w.writerow(["Indicador", "Valor"])
    w.writerow(["Rascunho", cp["rascunho"]])
    w.writerow(["Confirmados", cp["confirmado"]])
    w.writerow(["Expedidos", cp["expedido"]])
    w.writerow(["Cancelados", cp["cancelado"]])
    w.writerow(["Valor total expedido", dados["comercial"]["valor_total_expedido"]])
    w.writerow(["Clientes ativos", dados["comercial"]["clientes_ativos"]])
    w.writerow([])

    fr = dados["financeiro"]
    w.writerow(["Financeiro"])
    w.writerow(["Indicador", "Valor"])
    w.writerow(["A receber em aberto", fr["contas_a_receber"]["total_em_aberto"]])
    w.writerow(["A receber vencido", fr["contas_a_receber"]["total_vencido"]])
    w.writerow(["Recebido (total)", fr["contas_a_receber"]["total_baixado"]])
    w.writerow(["A pagar em aberto", fr["contas_a_pagar"]["total_em_aberto"]])
    w.writerow(["A pagar vencido", fr["contas_a_pagar"]["total_vencido"]])
    w.writerow(["Pago (total)", fr["contas_a_pagar"]["total_baixado"]])
    w.writerow(["Saldo projetado (aberto)", fr["saldo_projetado_em_aberto"]])
    w.writerow([])

    w.writerow([f"Fluxo de Caixa Projetado (calculado em {fluxo_caixa['hoje']})"])
    w.writerow(["Faixa", "Entradas previstas", "Saídas previstas", "Saldo líquido", "Saldo acumulado"])
    for b in fluxo_caixa["buckets"]:
        w.writerow([b["rotulo"], b["entradas_previstas"], b["saidas_previstas"], b["saldo_liquido"], b["saldo_acumulado"]])
    w.writerow([])

    # Fase 42 — bloco extra "no período", só quando o filtro foi usado.
    periodo = dados.get("periodo") or {"aplicado": False}
    if periodo["aplicado"]:
        w.writerow([f"No período — de {periodo['data_inicio'] or '(sem início)'} até {periodo['data_fim'] or '(sem fim)'}"])
        w.writerow(["Indicador", "Valor"])
        w.writerow(["Ordens concluídas", periodo["ordens_concluidas_no_periodo"]])
        w.writerow(["Lotes aprovados", periodo["lotes_aprovados_no_periodo"]])
        w.writerow(["Lotes reprovados", periodo["lotes_reprovados_no_periodo"]])
        w.writerow(["Pedidos expedidos", periodo["pedidos_expedidos_no_periodo"]])
        w.writerow(["Valor expedido", periodo["valor_expedido_no_periodo"]])
        w.writerow(["Valor recebido", periodo["valor_recebido_no_periodo"]])
        w.writerow(["Valor pago", periodo["valor_pago_no_periodo"]])

    return buffer.getvalue()


@bp.get("/dashboard/csv")
@requires_permission("relatorios", "visualizar")
def dashboard_csv():
    """Fase 19 — mesma exportação do Painel Gerencial da Fase 18, só que em
    CSV (separado por `;`, com BOM UTF-8) em vez de PDF — pra quem quer
    abrir os números numa planilha e continuar trabalhando neles, em vez
    de só ler um documento formatado. Export puro, mesma permissão
    reaproveitada, mesmo rótulo sintético de auditoria."""
    usuario_atual = g.usuario_atual
    conn = get_db()
    data_inicio = request.args.get("data_inicio") or None
    data_fim = request.args.get("data_fim") or None
    empresa_id = request.args.get("empresa_id", type=int)
    if empresa_id:
        _empresa_ou_404(conn, empresa_id)
    dados = _montar_dashboard(conn, data_inicio, data_fim, empresa_id)
    fluxo_caixa = _fluxo_caixa_projetado(conn, empresa_id)
    csv_texto = _gerar_csv_dashboard(dados, fluxo_caixa)

    audit.registrar(
        conn, tabela="painel_gerencial", registro_id=None, usuario_id=usuario_atual["id"],
        acao="painel_csv_gerado", valor_novo={"gerado_em": dados["gerado_em"]},
        ip=client_ip(), dispositivo=client_device(),
    )

    data_arquivo = datetime.datetime.utcnow().strftime("%Y-%m-%d")
    nome_arquivo = f"Painel-Gerencial-{data_arquivo}.csv"
    return Response(
        csv_texto.encode("utf-8"),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


# ============================================================
# Fase 45 — Exportação do Painel Gerencial em XLSX de verdade
# ============================================================
# A Fase 19 documentou de propósito que um .xlsx "de verdade" (com abas,
# formatação de célula e fórmulas — não só um CSV cru que o Excel abre)
# ficaria de fora por depender de uma biblioteca nova (`openpyxl`), que na
# época não estava disponível no ambiente onde aquela fase foi construída.
# Ela passou a estar disponível — esta fase entrega o que ficou pendente,
# reaproveitando exatamente a mesma `_montar_dashboard`/
# `_fluxo_caixa_projetado` da Fase 18/19 (o PDF, o CSV e o XLSX nunca podem
# divergir um do outro), a mesma permissão `relatorios.visualizar`, e o
# mesmo princípio de export puro (nunca altera nenhum dado de negócio).
_COR_CABECALHO_XLSX = "1F3A5F"  # mesmo azul do cabeçalho da tabela no PDF (Fase 18)
_FORMATO_MOEDA_XLSX = '"R$" #,##0.00'


def _gerar_xlsx_dashboard(dados, fluxo_caixa):
    wb = Workbook()

    fonte_titulo = Font(bold=True, size=14, color=_COR_CABECALHO_XLSX)
    fonte_subtitulo = Font(italic=True, size=9, color="5A6472")
    fonte_secao = Font(bold=True, size=11, color=_COR_CABECALHO_XLSX)
    fonte_cabecalho_tabela = Font(bold=True, color="FFFFFF")
    preenchimento_cabecalho = PatternFill("solid", fgColor=_COR_CABECALHO_XLSX)

    def _duas_colunas(ws, linha, rotulo, valor, moeda=False):
        ws.cell(row=linha, column=1, value=rotulo)
        celula_valor = ws.cell(row=linha, column=2, value=valor)
        if moeda:
            celula_valor.number_format = _FORMATO_MOEDA_XLSX
        return linha + 1

    def _titulo_secao(ws, linha, texto):
        celula = ws.cell(row=linha, column=1, value=texto)
        celula.font = fonte_secao
        return linha + 1

    # ---------- Aba 1: "Painel Gerencial" (os 5 blocos de sempre + período) ----------
    ws1 = wb.active
    ws1.title = "Painel Gerencial"
    ws1.cell(row=1, column=1, value="Alphafitus Laboratório Nutracêutico Ltda. — Painel Gerencial").font = fonte_titulo
    ws1.cell(row=2, column=1, value=f"Gerado em {dados['gerado_em']}").font = fonte_subtitulo
    linha = 3
    # Fase 52 — mesmo motivo do PDF/CSV: deixar explícito qual empresa os
    # números representam quando o filtro foi aplicado.
    empresa_filtrada = dados.get("empresa_filtrada")
    if empresa_filtrada:
        nome_empresa = empresa_filtrada["nome_fantasia"] or empresa_filtrada["razao_social"]
        ws1.cell(row=linha, column=1, value=f"Filtrando por: {nome_empresa}").font = fonte_subtitulo
        linha += 1
    linha += 1

    op = dados["producao"]["ordens_por_status"]
    linha = _titulo_secao(ws1, linha, "Produção")
    linha = _duas_colunas(ws1, linha, "Planejadas", op["planejada"])
    linha = _duas_colunas(ws1, linha, "Liberadas", op["liberada"])
    linha = _duas_colunas(ws1, linha, "Em produção", op["em_producao"])
    linha = _duas_colunas(ws1, linha, "Concluídas", op["concluida"])
    linha = _duas_colunas(ws1, linha, "Canceladas", op["cancelada"])
    linha = _duas_colunas(ws1, linha, "Lotes produzidos (total)", dados["producao"]["lotes_produzidos_total"])
    linha += 1

    lp = dados["qualidade"]["lotes_por_status"]
    ds = dados["qualidade"]["desvios_por_status"]
    taxa = dados["qualidade"]["taxa_aprovacao_lotes_pct"]
    linha = _titulo_secao(ws1, linha, "Qualidade")
    linha = _duas_colunas(ws1, linha, "Aprovados", lp["aprovado"] + lp["aprovado_com_ressalva"])
    linha = _duas_colunas(ws1, linha, "Reprovados", lp["reprovado"])
    linha = _duas_colunas(ws1, linha, "Em quarentena/análise", lp["quarentena"] + lp["em_analise"] + lp["aguardando_aprovacao"])
    linha = _duas_colunas(ws1, linha, "Taxa de aprovação (%)", taxa if taxa is not None else "—")
    linha = _duas_colunas(ws1, linha, "Análises aguardando resultado", dados["qualidade"]["analises_aguardando_resultado"])
    linha = _duas_colunas(ws1, linha, "Desvios em aberto", ds["aberto"])
    linha = _duas_colunas(ws1, linha, "Desvios em tratativa", ds["em_tratativa"])
    linha += 1

    est = dados["estoque"]
    linha = _titulo_secao(ws1, linha, "Estoque (WMS)")
    linha = _duas_colunas(ws1, linha, "Posições ativas", est["posicoes_ativas"])
    linha = _duas_colunas(ws1, linha, "Lotes vencidos com saldo", est["lotes_vencidos_com_saldo"])
    linha = _duas_colunas(ws1, linha, "A vencer em 30 dias", est["lotes_a_vencer_30_dias_com_saldo"])
    for tipo, saldo in est["saldo_total_por_tipo_item"].items():
        linha = _duas_colunas(ws1, linha, f"Saldo — {_ROTULOS_TIPO_ITEM.get(tipo, tipo)}", saldo)
    linha += 1

    cp = dados["comercial"]["pedidos_por_status"]
    linha = _titulo_secao(ws1, linha, "Comercial")
    linha = _duas_colunas(ws1, linha, "Rascunho", cp["rascunho"])
    linha = _duas_colunas(ws1, linha, "Confirmados", cp["confirmado"])
    linha = _duas_colunas(ws1, linha, "Expedidos", cp["expedido"])
    linha = _duas_colunas(ws1, linha, "Cancelados", cp["cancelado"])
    linha = _duas_colunas(ws1, linha, "Valor total expedido", dados["comercial"]["valor_total_expedido"], moeda=True)
    linha = _duas_colunas(ws1, linha, "Clientes ativos", dados["comercial"]["clientes_ativos"])
    linha += 1

    fr = dados["financeiro"]
    linha = _titulo_secao(ws1, linha, "Financeiro")
    linha = _duas_colunas(ws1, linha, "A receber em aberto", fr["contas_a_receber"]["total_em_aberto"], moeda=True)
    linha = _duas_colunas(ws1, linha, "  dos quais vencido", fr["contas_a_receber"]["total_vencido"], moeda=True)
    linha = _duas_colunas(ws1, linha, "Recebido (total)", fr["contas_a_receber"]["total_baixado"], moeda=True)
    linha = _duas_colunas(ws1, linha, "A pagar em aberto", fr["contas_a_pagar"]["total_em_aberto"], moeda=True)
    linha = _duas_colunas(ws1, linha, "  dos quais vencido", fr["contas_a_pagar"]["total_vencido"], moeda=True)
    linha = _duas_colunas(ws1, linha, "Pago (total)", fr["contas_a_pagar"]["total_baixado"], moeda=True)
    linha = _duas_colunas(ws1, linha, "Saldo projetado (aberto)", fr["saldo_projetado_em_aberto"], moeda=True)
    linha += 1

    # Fase 42 — bloco extra "no período", só quando o filtro foi usado.
    periodo = dados.get("periodo") or {"aplicado": False}
    if periodo["aplicado"]:
        linha = _titulo_secao(
            ws1, linha,
            f"No período — de {periodo['data_inicio'] or '(sem início)'} até {periodo['data_fim'] or '(sem fim)'}",
        )
        linha = _duas_colunas(ws1, linha, "Ordens concluídas", periodo["ordens_concluidas_no_periodo"])
        linha = _duas_colunas(ws1, linha, "Lotes aprovados", periodo["lotes_aprovados_no_periodo"])
        linha = _duas_colunas(ws1, linha, "Lotes reprovados", periodo["lotes_reprovados_no_periodo"])
        linha = _duas_colunas(ws1, linha, "Pedidos expedidos", periodo["pedidos_expedidos_no_periodo"])
        linha = _duas_colunas(ws1, linha, "Valor expedido", periodo["valor_expedido_no_periodo"], moeda=True)
        linha = _duas_colunas(ws1, linha, "Valor recebido", periodo["valor_recebido_no_periodo"], moeda=True)
        linha = _duas_colunas(ws1, linha, "Valor pago", periodo["valor_pago_no_periodo"], moeda=True)

    ws1.column_dimensions["A"].width = 34
    ws1.column_dimensions["B"].width = 18

    # ---------- Aba 2: "Fluxo de Caixa Projetado" (Fase 15) ----------
    # Tabela de verdade (cabeçalho colorido, moeda formatada) mais uma linha
    # de TOTAL calculada com uma fórmula `=SUM(...)` de dentro da própria
    # planilha (não um valor já somado em Python) — a diferença que a Fase
    # 19 apontou como faltando num CSV cru: aqui, se o usuário mudar um
    # valor na planilha, o total recalcula sozinho, do jeito que uma
    # planilha "de verdade" se espera que funcione.
    ws2 = wb.create_sheet("Fluxo de Caixa Projetado")
    cabecalho = ["Faixa", "Entradas previstas", "Saídas previstas", "Saldo líquido", "Saldo acumulado"]
    for coluna, texto in enumerate(cabecalho, start=1):
        celula = ws2.cell(row=1, column=coluna, value=texto)
        celula.font = fonte_cabecalho_tabela
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(horizontal="center")

    buckets = fluxo_caixa["buckets"]
    for indice, b in enumerate(buckets):
        linha_planilha = indice + 2
        ws2.cell(row=linha_planilha, column=1, value=b["rotulo"])
        for coluna, chave in ((2, "entradas_previstas"), (3, "saidas_previstas"), (4, "saldo_liquido"), (5, "saldo_acumulado")):
            celula = ws2.cell(row=linha_planilha, column=coluna, value=b[chave])
            celula.number_format = _FORMATO_MOEDA_XLSX

    linha_total = len(buckets) + 2
    ws2.cell(row=linha_total, column=1, value="Total").font = Font(bold=True)
    for coluna, letra in ((2, "B"), (3, "C"), (4, "D")):
        celula = ws2.cell(row=linha_total, column=coluna, value=f"=SUM({letra}2:{letra}{linha_total - 1})")
        celula.number_format = _FORMATO_MOEDA_XLSX
        celula.font = Font(bold=True)
    # "Saldo acumulado" é, por definição, o último valor da série (não uma
    # soma de todos os buckets) — copiamos o último em vez de somar tudo,
    # senão o total ficaria contábil/matematicamente errado.
    celula_acumulado_total = ws2.cell(row=linha_total, column=5, value=f"=E{linha_total - 1}")
    celula_acumulado_total.number_format = _FORMATO_MOEDA_XLSX
    celula_acumulado_total.font = Font(bold=True)

    for coluna, largura in enumerate((18, 20, 20, 18, 18), start=1):
        ws2.column_dimensions[get_column_letter(coluna)].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@bp.get("/dashboard/xlsx")
@requires_permission("relatorios", "visualizar")
def dashboard_xlsx():
    """Fase 45 — mesma exportação do Painel Gerencial das Fases 18 (PDF) e
    19 (CSV), agora também como um `.xlsx` de verdade (abas, formatação de
    célula e uma fórmula `=SUM(...)` de verdade na aba de Fluxo de Caixa)
    usando `openpyxl`. Export puro, mesma permissão reaproveitada, mesmo
    rótulo sintético de auditoria das duas exportações anteriores."""
    usuario_atual = g.usuario_atual
    conn = get_db()
    data_inicio = request.args.get("data_inicio") or None
    data_fim = request.args.get("data_fim") or None
    empresa_id = request.args.get("empresa_id", type=int)
    if empresa_id:
        _empresa_ou_404(conn, empresa_id)
    dados = _montar_dashboard(conn, data_inicio, data_fim, empresa_id)
    fluxo_caixa = _fluxo_caixa_projetado(conn, empresa_id)
    xlsx_bytes = _gerar_xlsx_dashboard(dados, fluxo_caixa)

    audit.registrar(
        conn, tabela="painel_gerencial", registro_id=None, usuario_id=usuario_atual["id"],
        acao="painel_xlsx_gerado", valor_novo={"gerado_em": dados["gerado_em"]},
        ip=client_ip(), dispositivo=client_device(),
    )

    data_arquivo = datetime.datetime.utcnow().strftime("%Y-%m-%d")
    nome_arquivo = f"Painel-Gerencial-{data_arquivo}.xlsx"
    return Response(
        xlsx_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


# =============================================================================
# Fase 150 — desempenho de visitas por vendedor do App de Vendas ("assim
# conseguimos monitorar todas as visitas e o tempo entre as visitas...
# avaliar desempenho por usuário do app", pedido do usuário). Deliberadamente
# em `relatorios.visualizar` (não `comercial.visualizar`, que o perfil
# "Vendedor" também tem) — mesmo raciocínio do resto deste arquivo: é uma
# visão que COMPARA pessoas entre si, não a operação de um vendedor sobre um
# cliente específico; um vendedor não deveria ver o desempenho dos colegas
# só por ter acesso ao app de vendas.
# =============================================================================
def _horas_entre(iso_inicio, iso_fim):
    inicio = datetime.datetime.strptime(iso_inicio[:19], "%Y-%m-%dT%H:%M:%S")
    fim = datetime.datetime.strptime(iso_fim[:19], "%Y-%m-%dT%H:%M:%S")
    return (fim - inicio).total_seconds() / 3600


@bp.get("/visitas-vendedores")
@requires_permission("relatorios", "visualizar")
def relatorio_visitas_vendedores():
    conn = get_db()
    data_inicio = request.args.get("data_inicio") or (datetime.datetime.utcnow() - datetime.timedelta(days=30)).strftime("%Y-%m-%d")
    data_fim = request.args.get("data_fim") or _hoje_iso_data()

    visitas = conn.execute(
        """
        SELECT vc.vendedor_id, u.nome AS vendedor_nome, vc.cliente_id, vc.chegada_em, vc.saida_em
        FROM visitas_clientes vc
        JOIN usuarios u ON u.id = vc.vendedor_id
        WHERE date(vc.chegada_em) BETWEEN ? AND ?
        ORDER BY vc.vendedor_id, vc.chegada_em
        """,
        (data_inicio, data_fim),
    ).fetchall()

    por_vendedor = {}
    for v in visitas:
        info = por_vendedor.setdefault(v["vendedor_id"], {
            "vendedor_id": v["vendedor_id"], "vendedor_nome": v["vendedor_nome"],
            "total_visitas": 0, "clientes_distintos": set(), "visitas_ainda_abertas": 0,
            "_chegadas": [], "_duracoes_horas": [],
        })
        info["total_visitas"] += 1
        info["clientes_distintos"].add(v["cliente_id"])
        info["_chegadas"].append(v["chegada_em"])
        if v["saida_em"]:
            info["_duracoes_horas"].append(_horas_entre(v["chegada_em"], v["saida_em"]))
        else:
            info["visitas_ainda_abertas"] += 1

    resultado = []
    for info in por_vendedor.values():
        chegadas = sorted(info.pop("_chegadas"))
        duracoes = info.pop("_duracoes_horas")
        info["clientes_distintos"] = len(info["clientes_distintos"])
        info["primeira_visita_em"] = chegadas[0]
        info["ultima_visita_em"] = chegadas[-1]
        # Tempo médio ENTRE visitas consecutivas (qualquer cliente) — "o
        # tempo entre as visitas", pedido do usuário — e tempo médio de
        # DURAÇÃO de cada visita (chegada→saída) como métrica extra.
        if len(chegadas) >= 2:
            intervalos = [_horas_entre(chegadas[i - 1], chegadas[i]) for i in range(1, len(chegadas))]
            info["media_horas_entre_visitas"] = round(sum(intervalos) / len(intervalos), 1)
        else:
            info["media_horas_entre_visitas"] = None
        info["media_horas_duracao_visita"] = round(sum(duracoes) / len(duracoes), 1) if duracoes else None
        resultado.append(info)

    resultado.sort(key=lambda r: r["total_visitas"], reverse=True)
    return jsonify({"data_inicio": data_inicio, "data_fim": data_fim, "vendedores": resultado})


@bp.get("/visitas-abertas")
@requires_permission("relatorios", "visualizar")
def relatorio_visitas_abertas():
    """Visitas que ainda não foram encerradas — inclusive as esquecidas há
    dias. Base pra `encerrar_visita_pelo_erp` abaixo."""
    conn = get_db()
    abertas = conn.execute(
        """
        SELECT vc.*, u.nome AS vendedor_nome, c.razao_social AS cliente_razao_social
        FROM visitas_clientes vc
        JOIN usuarios u ON u.id = vc.vendedor_id
        JOIN clientes c ON c.id = vc.cliente_id
        WHERE vc.saida_em IS NULL
        ORDER BY vc.chegada_em
        """,
    ).fetchall()
    return jsonify([dict(v) for v in abertas])


@bp.post("/visitas/<int:visita_id>/encerrar")
@requires_permission("relatorios", "visualizar")
def encerrar_visita_pelo_erp(visita_id):
    """"Esqueci de apontar minha saída" resolvido por quem administra,
    de dentro do ERP — pedido explícito do usuário: "ou então solicitar
    que seja feito por dentro do ERP ao usuário da empresa que tem acesso,
    bastando clicar". Nunca captura coordenada (quem clica não está no
    cliente) — só fecha o registro com o horário de agora, marcado como
    encerrado pelo ERP, não pelo próprio vendedor."""
    usuario_atual = g.usuario_atual
    conn = get_db()
    visita = conn.execute("SELECT * FROM visitas_clientes WHERE id = ?", (visita_id,)).fetchone()
    if visita is None:
        raise ApiError("Visita não encontrada.", status=404)
    if visita["saida_em"] is not None:
        raise ApiError("Esta visita já foi encerrada.", status=409)

    agora = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    conn.execute(
        "UPDATE visitas_clientes SET saida_em = ?, saida_registrada_por = ?, saida_encerrada_pelo_erp = 1 WHERE id = ?",
        (agora, usuario_atual["id"], visita_id),
    )
    audit.registrar(conn, tabela="visitas_clientes", registro_id=visita_id, usuario_id=usuario_atual["id"],
                     acao="visita_encerrada_pelo_erp", valor_novo={"vendedor_id": visita["vendedor_id"], "cliente_id": visita["cliente_id"]},
                     ip=client_ip(), dispositivo=client_device())

    atualizada = conn.execute("SELECT * FROM visitas_clientes WHERE id = ?", (visita_id,)).fetchone()
    return jsonify(dict(atualizada))


# =============================================================================
# Fase 150 — crédito pessoal do vendedor ("gordurinha"): visão de quem
# administra + transferência entre vendedores. Revisão do usuário: "o
# vendedor só pode usar seu crédito com seus clientes, onde o admin pode
# transferir essas verbas caso necessário para outro vendedor" — a
# transferência em si é só um par de lançamentos no MESMO ledger append-
# only que a geração automática usa (`creditos_vendedor_lancamentos`),
# nunca edita/apaga nada — mesmo raciocínio do resto do arquivo.
# =============================================================================
def _saldo_credito_vendedor(conn, vendedor_id):
    row = conn.execute(
        "SELECT COALESCE(SUM(CASE WHEN tipo = 'gerado' THEN valor ELSE -valor END), 0) AS saldo "
        "FROM creditos_vendedor_lancamentos WHERE vendedor_id = ?",
        (vendedor_id,),
    ).fetchone()
    return row["saldo"]


@bp.get("/creditos-vendedores")
@requires_permission("relatorios", "visualizar")
def relatorio_creditos_vendedores():
    conn = get_db()
    vendedores = conn.execute(
        """
        SELECT DISTINCT u.id, u.nome FROM usuarios u
        JOIN usuario_perfil up ON up.usuario_id = u.id
        JOIN perfil_permissao pp ON pp.perfil_id = up.perfil_id
        JOIN permissoes p ON p.id = pp.permissao_id
        WHERE p.modulo = 'vendas_app' AND p.acao = 'usar' AND u.status = 'ativo'
        ORDER BY u.nome
        """
    ).fetchall()
    return jsonify([{"vendedor_id": v["id"], "vendedor_nome": v["nome"], "saldo_disponivel": _saldo_credito_vendedor(conn, v["id"])} for v in vendedores])


@bp.post("/creditos-vendedor/transferir")
@requires_permission("relatorios", "visualizar")
def transferir_credito_vendedor():
    """Move saldo do crédito pessoal de um vendedor pra outro — grava como
    'utilizado' pra origem e 'gerado' pro destino, os dois com a MESMA
    observação linkando um ao outro, pra sempre dar pra rastrear a
    transferência olhando o extrato de qualquer um dos dois lados."""
    usuario_atual = g.usuario_atual
    conn = get_db()
    dados = request.get_json(silent=True) or {}
    vendedor_origem_id = dados.get("vendedor_origem_id")
    vendedor_destino_id = dados.get("vendedor_destino_id")
    valor = dados.get("valor")
    observacao = (dados.get("observacao") or "").strip()

    if not vendedor_origem_id or not vendedor_destino_id:
        raise ApiError("Informe vendedor_origem_id e vendedor_destino_id.", status=400)
    if vendedor_origem_id == vendedor_destino_id:
        raise ApiError("Origem e destino não podem ser o mesmo vendedor.", status=400)
    try:
        valor = float(valor)
    except (TypeError, ValueError):
        raise ApiError("Informe um valor numérico maior que zero.", status=400)
    if valor <= 0:
        raise ApiError("O valor a transferir precisa ser maior que zero.", status=400)

    for vendedor_id in (vendedor_origem_id, vendedor_destino_id):
        if not conn.execute("SELECT 1 FROM usuarios WHERE id = ? AND status = 'ativo'", (vendedor_id,)).fetchone():
            raise ApiError(f"Vendedor {vendedor_id} não encontrado ou inativo.", status=404)

    saldo_origem = _saldo_credito_vendedor(conn, vendedor_origem_id)
    if valor > saldo_origem + 0.0000001:
        raise ApiError(f"O vendedor de origem só tem R$ {saldo_origem:.2f} de crédito disponível.", status=400)

    texto_observacao = f"Transferência entre vendedores" + (f" — {observacao}" if observacao else "") + f" (feita por {usuario_atual['nome']})"
    conn.execute(
        "INSERT INTO creditos_vendedor_lancamentos (vendedor_id, tipo, valor, observacao, criado_por) VALUES (?, 'utilizado', ?, ?, ?)",
        (vendedor_origem_id, valor, texto_observacao, usuario_atual["id"]),
    )
    conn.execute(
        "INSERT INTO creditos_vendedor_lancamentos (vendedor_id, tipo, valor, observacao, criado_por) VALUES (?, 'gerado', ?, ?, ?)",
        (vendedor_destino_id, valor, texto_observacao, usuario_atual["id"]),
    )
    audit.registrar(conn, tabela="creditos_vendedor_lancamentos", registro_id=None, usuario_id=usuario_atual["id"],
                     acao="credito_vendedor_transferido",
                     valor_novo={"vendedor_origem_id": vendedor_origem_id, "vendedor_destino_id": vendedor_destino_id, "valor": valor, "observacao": observacao},
                     ip=client_ip(), dispositivo=client_device())

    return jsonify({
        "vendedor_origem_id": vendedor_origem_id, "saldo_origem_apos": _saldo_credito_vendedor(conn, vendedor_origem_id),
        "vendedor_destino_id": vendedor_destino_id, "saldo_destino_apos": _saldo_credito_vendedor(conn, vendedor_destino_id),
    })
